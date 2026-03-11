import os
from datetime import datetime
from sqlalchemy.orm import Session
from database import Invoice, LineItem
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from config import settings

def _header_style(cell, bg_color="1a1a2e"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        bottom=Side(style="medium", color="4a90d9")
    )

def _money(val):
    return round(float(val or 0), 2)

def generate_report(db: Session) -> str:
    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(settings.REPORTS_DIR, f"invoice_report_{timestamp}.xlsx")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Invoices ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoices"
    headers1 = ["ID", "Vendor", "Invoice No.", "Invoice Date", "Due Date",
                 "Total Amount", "Tax Amount", "Status", "Category", "File Path"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        _header_style(cell)
        ws1.column_dimensions[get_column_letter(col)].width = 18

    invoices = db.query(Invoice).all()
    for row, inv in enumerate(invoices, 2):
        ws1.cell(row=row, column=1, value=inv.id)
        ws1.cell(row=row, column=2, value=inv.vendor_name)
        ws1.cell(row=row, column=3, value=inv.invoice_number)
        ws1.cell(row=row, column=4, value=inv.invoice_date)
        ws1.cell(row=row, column=5, value=inv.due_date)
        ws1.cell(row=row, column=6, value=_money(inv.total_amount))
        ws1.cell(row=row, column=7, value=_money(inv.tax_amount))
        ws1.cell(row=row, column=8, value=inv.payment_status)
        ws1.cell(row=row, column=9, value=inv.category)
        ws1.cell(row=row, column=10, value=inv.file_path)
        if row % 2 == 0:
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = PatternFill("solid", fgColor="f0f4ff")

    # ── Sheet 2: Line Items ───────────────────────────────────────────
    ws2 = wb.create_sheet("Line Items")
    headers2 = ["Item ID", "Invoice ID", "Vendor", "Description", "Quantity", "Unit Price", "Item Total"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        _header_style(cell, "16213e")
        ws2.column_dimensions[get_column_letter(col)].width = 20

    line_items = db.query(LineItem).join(Invoice).all()
    for row, item in enumerate(line_items, 2):
        ws2.cell(row=row, column=1, value=item.id)
        ws2.cell(row=row, column=2, value=item.invoice_id)
        ws2.cell(row=row, column=3, value=item.invoice.vendor_name if item.invoice else "")
        ws2.cell(row=row, column=4, value=item.description)
        ws2.cell(row=row, column=5, value=item.quantity)
        ws2.cell(row=row, column=6, value=_money(item.unit_price))
        ws2.cell(row=row, column=7, value=_money(item.item_total))
        if row % 2 == 0:
            for col in range(1, 8):
                ws2.cell(row=row, column=col).fill = PatternFill("solid", fgColor="f0f4ff")

    # ── Sheet 3: Category Summary + Chart ────────────────────────────
    ws3 = wb.create_sheet("Category Summary")
    categories = ["Office Expenses", "Tools & Software", "Travel & Petrol", "Utilities", "Other"]
    headers3 = ["Category", "Total Spend", "Invoice Count"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        _header_style(cell, "0f3460")
        ws3.column_dimensions[get_column_letter(col)].width = 22

    for row, cat in enumerate(categories, 2):
        cat_invoices = [inv for inv in invoices if inv.category == cat]
        total = sum(_money(inv.total_amount) for inv in cat_invoices)
        ws3.cell(row=row, column=1, value=cat)
        ws3.cell(row=row, column=2, value=total)
        ws3.cell(row=row, column=3, value=len(cat_invoices))
        if row % 2 == 0:
            for col in range(1, 4):
                ws3.cell(row=row, column=col).fill = PatternFill("solid", fgColor="e8f4fd")

    # Bar Chart
    bar = BarChart()
    bar.title = "Spend by Category"
    bar.style = 10
    bar.y_axis.title = "Amount (₹)"
    bar.x_axis.title = "Category"
    bar.width = 20
    bar.height = 12
    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(categories)+1)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(categories)+1)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    ws3.add_chart(bar, "E2")

    # Pie Chart
    pie = PieChart()
    pie.title = "Expense Distribution"
    pie.style = 10
    pie.width = 18
    pie.height = 12
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws3.add_chart(pie, "E20")

    wb.save(output_path)
    print(f"Report saved: {output_path}")
    return output_path
